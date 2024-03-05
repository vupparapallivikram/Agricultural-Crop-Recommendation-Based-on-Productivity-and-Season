from django.db.models import Count
from django.db.models import Q
from django.shortcuts import render, redirect, get_object_or_404
import datetime
import openpyxl


# Create your views here.
from Remote_User.models import ClientRegister_Model,crop_details,crop_prediction,detection_ratio

def login(request):


    if request.method == "POST" and 'submit1' in request.POST:

        username = request.POST.get('username')
        password = request.POST.get('password')
        try:
            enter = ClientRegister_Model.objects.get(username=username,password=password)
            request.session["userid"] = enter.id

            return redirect('Add_DataSet_Details')
        except:
            pass

    return render(request,'RUser/login.html')

def Add_DataSet_Details(request):
    if "GET" == request.method:
        return render(request, 'RUser/Add_DataSet_Details.html', {})
    else:
        excel_file = request.FILES["excel_file"]
        # you may put validations here to check extension or file size
        wb = openpyxl.load_workbook(excel_file)
        # getting all sheets
        sheets = wb.sheetnames
        print(sheets)
        # getting a particular sheet
        worksheet = wb["Sheet1"]
        print(worksheet)
        # getting active sheet
        active_sheet = wb.active
        print(active_sheet)
        # reading a cell
        print(worksheet["A1"].value)
        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
                print(cell.value)
            excel_data.append(row_data)
            crop_details.objects.all().delete()
    for r in range(1, active_sheet.max_row+1):
        crop_details.objects.create(
        State_Name= active_sheet.cell(r, 1).value,
        District_Name= active_sheet.cell(r, 2).value,
        Crop_Year= active_sheet.cell(r, 3).value,
        Season= active_sheet.cell(r, 4).value,
        names= active_sheet.cell(r, 5).value,
        Area= active_sheet.cell(r, 6).value,
        Production= active_sheet.cell(r, 7).value,

        )

    return render(request, 'RUser/Add_DataSet_Details.html', {"excel_data": excel_data})


def Register1(request):

    if request.method == "POST":
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')
        phoneno = request.POST.get('phoneno')
        country = request.POST.get('country')
        state = request.POST.get('state')
        city = request.POST.get('city')
        ClientRegister_Model.objects.create(username=username, email=email, password=password, phoneno=phoneno,
                                            country=country, state=state, city=city)

        return render(request, 'RUser/Register1.html')
    else:
        return render(request,'RUser/Register1.html')

def ViewYourProfile(request):
    userid = request.session['userid']
    obj = ClientRegister_Model.objects.get(id= userid)
    return render(request,'RUser/ViewYourProfile.html',{'object':obj})


def Predict_Crop_Yiled_OnDataSets(request):
        expense = 0
        kg_price=0
        if request.method == "POST":

            State = request.POST.get('State')
            District = request.POST.get('District')
            Year = request.POST.get('Year')
            Season = request.POST.get('Season')
            cname=request.POST.get('cname')
            area = request.POST.get('area')
            production = request.POST.get('production')
            area1=int(area)
            production1=int(production)

            if area1<10:
                expense=15000
            elif area1<50 and area1>10:
                expense=70000
            elif area1 < 100 and area1 > 50:
                expense = 100000
            elif area1 < 250 and area1 > 100:
                expense = 150000
                print(expense)
            elif area1 < 500 and area1 > 250:
                expense = 200000
            else:
                expense=300000

            if  cname=="Dry ginger":
                kg_price=100
            elif cname == "Sugarcane":
                 kg_price = 50
            elif cname == "Sweet potato":
                 kg_price = 40
            elif cname == "Sugarcane":
                kg_price = 50
            elif cname == "Rice":
                kg_price = 50
            elif cname == "Banana":
                kg_price = 70
                print(kg_price)
            elif cname == "Black pepper":
                kg_price = 1170
            elif cname == "Coconut":
                kg_price = 25
            elif cname == "Dry chillies":
                kg_price = 400
            elif cname == "Grapes":
                kg_price = 50
            elif cname == "Groundnut":
                kg_price = 170
            elif cname == "Horse-gram":
                kg_price = 70
            elif cname == "Jowar":
                kg_price = 80
            elif cname == "Maize":
                 kg_price = 50
            elif cname == "Moong_Green Gram":
                 kg_price = 40
            elif cname == "Onion":
                kg_price = 90
            elif cname == "Ragi":
                kg_price = 70
            elif cname == "Small millets":
                kg_price = 120
            elif cname == "Soyabean":
                kg_price = 170
            elif cname == "Urad":
                kg_price = 125
            elif cname == "Bajra":
                kg_price = 400
            elif cname == "Turmeric":
                kg_price = 1250
            elif cname == "Potato":
                kg_price = 50
            elif cname == "Wheat":
                kg_price = 90
            elif cname == "Coriander":
                kg_price = 280
            elif cname == "Arecanut":
                kg_price = 180


            yield1=(production1*(kg_price))-int(expense)

            prod=production1/area1

            crop_prediction.objects.create(State_Name=State,District_Name=District,Crop_Year=Year,Season=Season,names=cname,Area=area,Production=production,Yield_Prediction=yield1,Production_Prediction=prod)


            return render(request, 'RUser/Predict_Crop_Yiled_OnDataSets.html',{'objs':yield1,'objs1':prod})
        return render(request, 'RUser/Predict_Crop_Yiled_OnDataSets.html')


def ratings(request,pk):
    vott1, vott, neg = 0, 0, 0
    objs = crop_details.objects.get(id=pk)
    unid = objs.id
    vot_count = crop_details.objects.all().filter(id=unid)
    for t in vot_count:
        vott = t.ratings
        vott1 = vott + 1
        obj = get_object_or_404(crop_details, id=unid)
        obj.ratings = vott1
        obj.save(update_fields=["ratings"])
        return redirect('Add_DataSet_Details')

    return render(request,'RUser/ratings.html',{'objs':vott1})



